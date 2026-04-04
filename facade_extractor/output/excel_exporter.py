"""
excel_exporter.py
─────────────────
Writes ExtractionResult objects to an Excel workbook using openpyxl.

Layout:
  • One sheet per drawing file (named by file stem)
  • Columns: ID | Parameter Name | Extracted Value | Unit | Confidence |
             Extraction Method | Source Layer | Spec Value | Spec Check |
             Delta | Notes
  • Cell fill colours:
      Green  #90FF90 → MATCH,    confidence ≥ 0.75
      Yellow #FFFF90 → NO_SPEC,  confidence ≥ 0.75
      Orange #FFB347 → any extracted, confidence < 0.75
      Red    #FF9090 → CONFLICT
      Grey   #C0C0C0 → NOT_FOUND
  • Header: dark navy background, white bold text
  • Summary sheet with extraction status pie chart + confidence bar chart
"""

from __future__ import annotations

from pathlib import Path
from typing import Union

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side,
    )
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.chart.series import DataPoint
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from output.result_builder import ExtractionResult

# ─────────────────────────────────────────────────────────────────────────────
# Style constants
# ─────────────────────────────────────────────────────────────────────────────

_FILL_GREEN  = "90FF90"
_FILL_YELLOW = "FFFF90"
_FILL_ORANGE = "FFB347"
_FILL_RED    = "FF9090"
_FILL_GREY   = "C0C0C0"
_FILL_HEADER = "1F3864"   # dark navy

_COLUMNS = [
    "ID",
    "Parameter Name",
    "Extracted Value",
    "Unit",
    "Confidence",
    "Extraction Method",
    "Source Layer",
    "Spec Value",
    "Spec Check",
    "Delta",
    "Notes",
]

_COL_WIDTHS = [6, 30, 16, 6, 12, 20, 20, 12, 12, 10, 30]


def _fill(hex_color: str) -> "PatternFill":
    return PatternFill(fill_type="solid", fgColor=hex_color)


def _row_fill(param: dict) -> str:
    spec = param.get("spec_check", {})
    status = spec.get("result", "NO_SPEC")
    value  = param.get("value")
    conf   = param.get("confidence", 0.0) or 0.0

    if value is None or param.get("extraction_method") == "NOT_FOUND":
        return _FILL_GREY
    if status == "CONFLICT":
        return _FILL_RED
    if conf < 0.75:
        return _FILL_ORANGE
    if status == "MATCH":
        return _FILL_GREEN
    return _FILL_YELLOW   # NO_SPEC, high confidence


# ─────────────────────────────────────────────────────────────────────────────
# Main export
# ─────────────────────────────────────────────────────────────────────────────

def export_excel(
    results: list[ExtractionResult],
    output_dir: Union[str, Path],
    filename: str = "facade_extraction.xlsx",
) -> Path:
    """
    Write all results to one Excel workbook.

    Parameters
    ----------
    results    : list of ExtractionResult (one per file processed)
    output_dir : destination directory
    filename   : workbook filename

    Returns
    -------
    Path to written .xlsx file.
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl is required for Excel export. pip install openpyxl")

    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    out_path = output_dir / filename

    wb = Workbook()
    # Remove default sheet
    if wb.active:
        wb.remove(wb.active)

    # ── One sheet per result ──────────────────────────────────────────────
    summary_data: list[dict] = []

    for result in results:
        sheet_name = _safe_sheet_name(Path(result.input_file).stem)
        ws = wb.create_sheet(title=sheet_name)
        _write_result_sheet(ws, result)

        s = result.extraction_summary
        summary_data.append({
            "file":           result.input_file,
            "extracted":      s.get("parameters_extracted", 0),
            "not_found":      s.get("parameters_not_found", 0),
            "conflicts":      s.get("conflicts_with_spec", 0),
            "avg_confidence": s.get("average_confidence", 0.0),
        })

    # ── Summary sheet ─────────────────────────────────────────────────────
    if summary_data:
        _write_summary_sheet(wb, summary_data, results)

    wb.save(str(out_path))
    return out_path


# ─────────────────────────────────────────────────────────────────────────────
# Per-drawing sheet writer
# ─────────────────────────────────────────────────────────────────────────────

def _write_result_sheet(ws: Any, result: ExtractionResult) -> None:
    # ── Metadata header rows ──────────────────────────────────────────────
    meta = result.sheet_metadata
    ws.append(["File", result.input_file, "", "Pipeline", result.processing_pipeline])
    ws.append(["Sheet", meta.get("sheet_title", ""), "", "Scale", meta.get("scale", "")])
    ws.append(["Type", meta.get("sheet_type", ""), "", "Rev", meta.get("revision", "")])
    ws.append([])   # blank row

    # ── Column headers ────────────────────────────────────────────────────
    header_row = ws.max_row + 1
    ws.append(_COLUMNS)

    header_font  = Font(color="FFFFFF", bold=True)
    header_fill  = _fill(_FILL_HEADER)
    header_align = Alignment(horizontal="center", vertical="center")
    header_border = Border(
        bottom=Side(style="thin", color="FFFFFF")
    )

    for col_idx, col_name in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        cell.border    = header_border
        ws.column_dimensions[get_column_letter(col_idx)].width = _COL_WIDTHS[col_idx - 1]

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    ws.auto_filter.ref = (
        f"A{header_row}:{get_column_letter(len(_COLUMNS))}{header_row}"
    )

    # ── Data rows ─────────────────────────────────────────────────────────
    for param in result.parameters:
        spec   = param.get("spec_check", {})
        value  = param.get("value")
        row = [
            param.get("id", ""),
            param.get("name", ""),
            round(value, 2) if value is not None else "—",
            param.get("unit", "mm"),
            f"{param.get('confidence', 0.0):.0%}" if param.get("confidence") else "—",
            param.get("extraction_method", ""),
            param.get("source_layer", ""),
            spec.get("spec_value", "—"),
            spec.get("result", "NO_SPEC"),
            round(spec["delta"], 2) if spec.get("delta") is not None else "—",
            param.get("notes", ""),
        ]
        ws.append(row)

        # Apply row fill
        fill_color = _row_fill(param)
        row_fill = _fill(fill_color)
        for col_idx in range(1, len(_COLUMNS) + 1):
            ws.cell(row=ws.max_row, column=col_idx).fill = row_fill

    # Alignment
    for row in ws.iter_rows(min_row=header_row + 1):
        for cell in row:
            cell.alignment = Alignment(vertical="center")


# ─────────────────────────────────────────────────────────────────────────────
# Summary sheet
# ─────────────────────────────────────────────────────────────────────────────

def _write_summary_sheet(
    wb: Any,
    summary_data: list[dict],
    results: list[ExtractionResult],
) -> None:
    ws = wb.create_sheet(title="Summary", index=0)

    ws.append(["FACADE PARAMETER EXTRACTION — SUMMARY"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])

    headers = ["File", "Extracted", "Not Found", "Conflicts", "Avg Confidence"]
    ws.append(headers)
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = _fill(_FILL_HEADER)

    data_start_row = ws.max_row + 1
    totals = {"extracted": 0, "not_found": 0, "conflicts": 0, "avg_conf_sum": 0.0}

    for s in summary_data:
        ws.append([
            s["file"],
            s["extracted"],
            s["not_found"],
            s["conflicts"],
            f"{s['avg_confidence']:.0%}",
        ])
        totals["extracted"]  += s["extracted"]
        totals["not_found"]  += s["not_found"]
        totals["conflicts"]  += s["conflicts"]
        totals["avg_conf_sum"] += s["avg_confidence"]

    n = len(summary_data) or 1
    ws.append([
        "TOTAL / AVG",
        totals["extracted"],
        totals["not_found"],
        totals["conflicts"],
        f"{totals['avg_conf_sum'] / n:.0%}",
    ])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    ws.column_dimensions["A"].width = 40
    for col in ["B", "C", "D", "E"]:
        ws.column_dimensions[col].width = 14

    # ── Pie chart: extraction status (first result only for simplicity) ───
    if results:
        _add_status_pie(ws, results[0])


def _add_status_pie(ws: Any, result: ExtractionResult) -> None:
    """Add a small pie chart showing extracted / not_found / conflicts."""
    s = result.extraction_summary
    stats = [
        ("Extracted",  s.get("parameters_extracted", 0)),
        ("Not Found",  s.get("parameters_not_found", 0)),
        ("Conflicts",  s.get("conflicts_with_spec",  0)),
    ]
    # Write data to hidden area
    start_row = 20
    for i, (label, val) in enumerate(stats):
        ws.cell(row=start_row + i, column=7, value=label)
        ws.cell(row=start_row + i, column=8, value=val)

    pie = PieChart()
    pie.title = "Extraction Status"
    pie.style = 10
    labels = Reference(ws, min_col=7, min_row=start_row, max_row=start_row + 2)
    data   = Reference(ws, min_col=8, min_row=start_row, max_row=start_row + 2)
    pie.add_data(data)
    pie.set_categories(labels)
    pie.width  = 12
    pie.height = 10
    ws.add_chart(pie, "G2")


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _safe_sheet_name(name: str, max_len: int = 31) -> str:
    """Excel sheet names: max 31 chars, no special chars."""
    import re
    safe = re.sub(r"[\\/:*?\[\]]", "_", name)
    return safe[:max_len]


# Allow Any type hint without runtime import
from typing import Any
